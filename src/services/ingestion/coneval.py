"""CONEVAL municipal poverty data ingestion.

Parses the CONEVAL "Concentrado, indicadores de pobreza 2020" XLSX
and extracts:
  - extreme_poverty (29): % poblacion en pobreza extrema (2020)
  - overcrowding_coneval (alt): % viviendas con hacinamiento (CONEVAL carencias)

The CONEVAL Excel has a complex merged-cell header structure:
  Row 4: Level-1 headers ("Pobreza extrema" at col 17)
  Row 5: Level-2 headers (Porcentaje 2010/2015/2020 at cols 17-22)
  Row 8+: Data rows

Data source: CONEVAL "Concentrado indicadores de pobreza 2020"
Download from: https://www.coneval.org.mx/Medicion/Paginas/Pobreza-municipal.aspx
Place in: Engine/data/coneval/
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from src.config import DATA_DIR

logger = logging.getLogger(__name__)

SOURCE_NAME = "CONEVAL"
CACHE_DIR = DATA_DIR / "coneval"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# CONEVAL "Concentrado municipal" sheet column positions (0-indexed)
# These are fixed positions from the 2020 CONEVAL Excel format.
_CONCENTRADO_COLS = {
    "cve_ent": 1,
    "entidad": 2,
    "cve_mun": 3,
    "municipio": 4,
    "poblacion_2020": 7,
    # "Pobreza extrema" group starts at col 17 (0-indexed)
    # Sub-columns: Porcentaje 2010, 2015, 2020, Personas 2010, 2015, 2020
    "pobreza_extrema_pct_2020": 17 + 2,  # col 19
    # "Carencia por calidad y espacios de la vivienda" starts at col 83
    "hacinamiento_pct_2020": 83 + 2,  # col 85
}

INDICATOR_MAP = {
    "extreme_poverty": {
        "column": "pobreza_extrema_pct_2020",
        "description": "% poblacion en pobreza extrema (CONEVAL 2020)",
        "unit": "%",
    },
}


def _find_xlsx_files(directory: Path) -> list[Path]:
    if not directory.exists():
        return []
    files = sorted(
        [f for f in directory.iterdir() if f.suffix.lower() == ".xlsx"],
        key=lambda f: (
            # Prefer files with "Concentrado" in name
            0 if "concentrado" in f.name.lower() else 1,
            # Then by mtime (newest first)
            -f.stat().st_mtime,
        ),
    )
    return files


def parse_coneval_data(
    *,
    data_path: Path | None = None,
) -> dict[str, dict[str, float]]:
    """Parse CONEVAL XLSX and return indicator values per municipality.

    Returns: {indicator_id: {municipio_code: value}}
    """
    import openpyxl

    if data_path and data_path.exists():
        xlsx_path = data_path
    else:
        xlsx_files = _find_xlsx_files(CACHE_DIR)
        if not xlsx_files:
            logger.info(f"{SOURCE_NAME}: no XLSX/CSV found in {CACHE_DIR}")
            return {}
        xlsx_path = xlsx_files[0]

    if not xlsx_path.exists():
        return {}

    logger.info(f"{SOURCE_NAME}: loading {xlsx_path.name}")

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    except Exception as exc:
        logger.warning(f"{SOURCE_NAME}: cannot open {xlsx_path.name}: {exc}")
        return {}

    sheet = wb[wb.sheetnames[0]]
    logger.info(f"{SOURCE_NAME}: sheet = {wb.sheetnames[0]}")

    result: dict[str, dict[str, float]] = {
        ind_id: {} for ind_id in INDICATOR_MAP
    }

    data_started = False
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i < 8:
            continue

        row_vals = list(row)
        cve_ent = _cell_str(row_vals, _CONCENTRADO_COLS["cve_ent"])
        cve_mun = _cell_str(row_vals, _CONCENTRADO_COLS["cve_mun"])

        if not cve_ent or not cve_mun:
            continue

        cve_ent = cve_ent.zfill(2)
        cve_mun = cve_mun.zfill(3)
        muni_code = cve_ent + cve_mun

        for ind_id, config in INDICATOR_MAP.items():
            col_idx = _CONCENTRADO_COLS[config["column"]]
            val = _cell_float(row_vals, col_idx)
            if val is not None and val >= 0:
                result[ind_id][muni_code] = round(val, 2)

    wb.close()

    logger.info(
        f"{SOURCE_NAME}: parsed {len(result.get('extreme_poverty', {}))} municipalities"
    )
    return result


def get_state_aggregates(
    coneval_data: dict[str, dict[str, float]],
    indicator_id: str,
) -> dict[str, float]:
    """Aggregate municipal values to state-level averages."""
    values = coneval_data.get(indicator_id, {})
    state_sums: dict[str, float] = {}
    state_counts: dict[str, int] = {}

    for muni_code, val in values.items():
        state = muni_code[:2]
        state_sums[state] = state_sums.get(state, 0.0) + val
        state_counts[state] = state_counts.get(state, 0) + 1

    return {
        state: round(state_sums[state] / state_counts[state], 2)
        for state in state_sums
    }


def _cell_str(row: list[Any], idx: int) -> str:
    if idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def _cell_float(row: list[Any], idx: int) -> float | None:
    if idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    try:
        return float(str(val).replace(",", "").replace("%", ""))
    except (ValueError, TypeError):
        return None


def check_available() -> bool:
    xlsx_files = _find_xlsx_files(CACHE_DIR)
    return len(xlsx_files) > 0
