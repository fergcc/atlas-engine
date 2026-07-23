"""CONAGUA EAM/SINA water-stress ingestion — real bulk-file replacement for
the hardcoded conagua.py fallback.

CONAGUA does not publish a documented API (per the Atlas methodology's
connector matrix: "No documented API"). "Estadísticas del Agua en México"
(EAM) and SINA publish state/basin water tables as downloadable Excel/CSV
annexes, refreshed annually. This module parses whatever such file a human
has placed in Engine/data/conagua_eam/, using flexible header-keyword
matching (state name + %/value column) since the exact column layout shifts
between EAM editions — mirrors the pattern in censo2020.py/coneval.py.

Download from:
  https://www.gob.mx/conagua/acciones-y-programas/estadisticas-del-agua-en-mexico
  https://sina.conagua.gob.mx/
Place the annex file (XLSX or CSV) in: Engine/data/conagua_eam/

If no file is present (or no matching column is found), every function here
returns {} rather than raising — callers (territorial.py) must fall back to
conagua.py's hardcoded FALLBACK_MX in that case, and must tag that fallback
as data_quality="synthetic", not "real" (mislabeling the static 2023 PDF
values as "real" was the original bug this module exists to fix).
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path

from src.config import DATA_DIR

logger = logging.getLogger(__name__)

SOURCE_NAME = "CONAGUA_EAM"
CACHE_DIR = DATA_DIR / "conagua_eam"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# EAM/SINA tables key rows by state name, not INEGI code.
_STATE_NAME_TO_CODE: dict[str, str] = {
    "aguascalientes": "01", "baja california": "02", "baja california sur": "03",
    "campeche": "04", "coahuila": "05", "coahuila de zaragoza": "05",
    "colima": "06", "chiapas": "07", "chihuahua": "08",
    "ciudad de mexico": "09", "distrito federal": "09",
    "durango": "10", "guanajuato": "11", "guerrero": "12", "hidalgo": "13",
    "jalisco": "14", "mexico": "15", "estado de mexico": "15",
    "michoacan": "16", "michoacan de ocampo": "16", "morelos": "17",
    "nayarit": "18", "nuevo leon": "19", "oaxaca": "20", "puebla": "21",
    "queretaro": "22", "quintana roo": "23", "san luis potosi": "24",
    "sinaloa": "25", "sonora": "26", "tabasco": "27", "tamaulipas": "28",
    "tlaxcala": "29", "veracruz": "30", "veracruz de ignacio de la llave": "30",
    "yucatan": "31", "zacatecas": "32",
}

_INDICATOR_KEYWORDS: dict[str, list[str]] = {
    "water_stress": ["estres", "presion", "grado de presion", "extraccion"],
    "water_consumption_intensity": ["consumo", "uso consuntivo", "volumen concesionado"],
}


def _find_files(directory: Path) -> list[Path]:
    if not directory.exists():
        return []
    return sorted(
        [f for f in directory.iterdir() if f.suffix.lower() in (".xlsx", ".csv")],
        key=lambda f: -f.stat().st_mtime,
    )


def _normalize_state(raw: str | None) -> str | None:
    if not raw:
        return None
    key = str(raw).strip().lower()
    for prefix in ("estado de ", "edo. de ", "edo de "):
        if key.startswith(prefix):
            key = key[len(prefix):]
    return _STATE_NAME_TO_CODE.get(key)


def _empty_result() -> dict[str, dict[str, float]]:
    return {ind_id: {} for ind_id in _INDICATOR_KEYWORDS}


def _parse_xlsx(path: Path) -> dict[str, dict[str, float]]:
    import openpyxl

    result = _empty_result()
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning(f"{SOURCE_NAME}: cannot open {path.name}: {exc}")
        return result

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        state_col = next((i for i, h in enumerate(header) if "entidad" in h or "estado" in h), None)
        if state_col is None:
            continue

        for ind_id, keywords in _INDICATOR_KEYWORDS.items():
            value_col = next(
                (i for i, h in enumerate(header) if any(kw in h for kw in keywords)), None
            )
            if value_col is None:
                continue
            for row in rows[1:]:
                if state_col >= len(row) or value_col >= len(row):
                    continue
                code = _normalize_state(row[state_col])
                if not code:
                    continue
                try:
                    val = float(str(row[value_col]).replace(",", "").replace("%", ""))
                except (ValueError, TypeError):
                    continue
                result[ind_id][code] = round(val, 2)

    wb.close()
    return result


def _parse_csv(path: Path) -> dict[str, dict[str, float]]:
    result = _empty_result()
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, encoding=encoding, newline="") as fh:
                reader = csv.reader(fh)
                header = [c.strip().lower() for c in next(reader)]
                state_col = next(
                    (i for i, h in enumerate(header) if "entidad" in h or "estado" in h), None
                )
                if state_col is None:
                    return result
                value_cols = {
                    ind_id: next(
                        (i for i, h in enumerate(header) if any(kw in h for kw in kws)), None
                    )
                    for ind_id, kws in _INDICATOR_KEYWORDS.items()
                }
                for row in reader:
                    if state_col >= len(row):
                        continue
                    code = _normalize_state(row[state_col])
                    if not code:
                        continue
                    for ind_id, col in value_cols.items():
                        if col is None or col >= len(row):
                            continue
                        try:
                            val = float(row[col].replace(",", "").replace("%", ""))
                        except (ValueError, TypeError):
                            continue
                        result[ind_id][code] = round(val, 2)
            return result
        except (UnicodeDecodeError, UnicodeError):
            continue
    return result


def parse_conagua_eam_data(*, data_path: Path | None = None) -> dict[str, dict[str, float]]:
    """Parse a CONAGUA EAM/SINA state water annex placed in CACHE_DIR.

    Returns {indicator_id: {state_code: value}} — an indicator maps to {} if
    no matching column was found in any candidate file. Never raises.
    """
    candidates = [data_path] if data_path and data_path.exists() else _find_files(CACHE_DIR)

    if not candidates:
        logger.info(
            f"{SOURCE_NAME}: no XLSX/CSV found in {CACHE_DIR}. Download the state "
            f"water table from CONAGUA's 'Estadisticas del Agua en Mexico' "
            f"(https://www.gob.mx/conagua/acciones-y-programas/estadisticas-del-agua-en-mexico) "
            f"or SINA (https://sina.conagua.gob.mx/) and place it in {CACHE_DIR}/"
        )
        return _empty_result()

    for path in candidates:
        result = _parse_xlsx(path) if path.suffix.lower() == ".xlsx" else _parse_csv(path)
        if any(result.values()):
            logger.info(
                f"{SOURCE_NAME}: parsed {path.name} — "
                + ", ".join(f"{k}: {len(v)} states" for k, v in result.items())
            )
            return result

    return _empty_result()


def check_available() -> bool:
    return len(_find_files(CACHE_DIR)) > 0
