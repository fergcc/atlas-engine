"""INEGI ENOE — Encuesta Nacional de Ocupación y Empleo.

Extracts state-level employment indicators from INEGI's ENOE tabulados:
  - employed_population (15): % población ocupada (PEA ocupada / PEA)
  - female_employment (16): % mujeres ocupadas en edad reproductiva
  - hours_worked (20): horas promedio trabajadas por semana
  - remuneration_level (19): ingreso promedio mensual (pesos)

Data source: INEGI ENOE Indicadores Estratégicos (trimestral)
Download from: https://www.inegi.org.mx/programas/enoe/15ymas/#tabulados
  → Buscar "Indicadores estratégicos" → descargar el ZIP más reciente
  → Extraer los XLSX en Engine/data/enoe/
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from src.config import DATA_DIR

logger = logging.getLogger(__name__)

SOURCE_NAME = "ENOE"
CACHE_DIR = DATA_DIR / "enoe"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# ENOE indicator columns in the Indicadores Estratégicos Excel format
# These Excel files have multiple sheets; we want the state-level summary.
# Typical file: "enei_2026_trim1.xlsx" with sheets like "Nacional", "Entidad Federativa"

INDICATOR_MAP = {
    "employed_population": {
        "description": "% población ocupada (PEA ocupada / PEA total, 15+ años)",
        "unit": "%",
    },
    "female_employment": {
        "description": "% mujeres ocupadas (15-49 años / total mujeres PEA)",
        "unit": "%",
    },
    "hours_worked": {
        "description": "Horas promedio trabajadas por semana",
        "unit": "horas/semana",
    },
    "remuneration_level": {
        "description": "Ingreso promedio mensual por trabajo",
        "unit": "MXN/mes",
    },
}


def _find_xlsx_files(directory: Path) -> list[Path]:
    if not directory.exists():
        return []
    files = sorted(
        [f for f in directory.iterdir() if f.suffix.lower() == ".xlsx"],
        key=lambda f: -f.stat().st_mtime,
    )
    return files


def parse_enoe_data(
    *,
    data_path: Path | None = None,
) -> dict[str, dict[str, float]]:
    """Parse ENOE Excel files and return indicator values per state.

    Returns: {indicator_id: {state_code: value}}

    If no XLSX files are found, logs download instructions.
    """
    import openpyxl

    if data_path and data_path.exists():
        xlsx_path = data_path
    else:
        xlsx_files = _find_xlsx_files(CACHE_DIR)
        if not xlsx_files:
            logger.info(
                f"{SOURCE_NAME}: no XLSX found in {CACHE_DIR}. "
                f"Download ENOE Indicadores Estratégicos from "
                f"https://www.inegi.org.mx/programas/enoe/15ymas/#tabulados "
                f"and place the extracted XLSX files in {CACHE_DIR}/"
            )
            return {}
        xlsx_path = xlsx_files[0]

    logger.info(f"{SOURCE_NAME}: parsing {xlsx_path.name}")

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning(f"{SOURCE_NAME}: cannot open {xlsx_path.name}: {exc}")
        return {}

    # Find the state-level sheet
    sheet_name = None
    for name in wb.sheetnames:
        if any(kw in name.lower() for kw in ["entidad", "estado", "state"]):
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]

    ws = wb[sheet_name]
    logger.info(f"{SOURCE_NAME}: reading sheet '{sheet_name}'")

    result: dict[str, dict[str, float]] = {
        ind_id: {} for ind_id in INDICATOR_MAP
    }

    # Find header row (search for "Entidad" or "Aguascalientes" or numeric entity codes)
    headers: dict[str, int] = {}
    data_start = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_vals = [str(c) if c is not None else "" for c in row]
        row_text = " ".join(row_vals).upper()
        if "ENTIDAD" in row_text or "AGUASCALIENTES" in row_text or "INDICADOR" in row_text:
            for j, val in enumerate(row):
                val_str = str(val).strip() if val else ""
                if val_str:
                    headers[val_str.lower()] = j
            data_start = i + 1
            break

    if not headers:
        # Try row 3-6 as headers (typical ENOE format)
        for i in range(3, min(8, ws.max_row or 10)):
            row_vals = []
            for row in enumerate(ws.iter_rows(min_row=i, max_row=i, values_only=True)):
                row_vals = [str(c) if c is not None else "" for c in row]
                break
            if any("entidad" in v.lower() for v in row_vals) or any(
                v.strip().isdigit() and len(v.strip()) == 2 for v in row_vals if v.strip()
            ):
                for j, val in enumerate(row_vals):
                    if val:
                        headers[val.lower()] = j
                data_start = i + 1
                break

    if not headers:
        logger.warning(f"{SOURCE_NAME}: could not find header row in {xlsx_path.name}")
        wb.close()
        return result

    logger.info(f"{SOURCE_NAME}: headers found at row {data_start}: {list(headers.keys())[:6]}")

    # Parse data rows
    for i, row in enumerate(ws.iter_rows(min_row=data_start or 1, values_only=True)):
        row_vals = [str(c) if c is not None else "" for c in row]

        # Find entity code
        ent_code = _find_entity_code(row_vals, headers)
        if not ent_code:
            # Try first column
            first_col = row_vals[0] if row_vals else ""
            if first_col.strip().isdigit() and len(first_col.strip()) == 2:
                ent_code = first_col.strip().zfill(2)
            else:
                ent_code = _resolve_entity_name_from_row(row_vals)
        if not ent_code:
            continue

        # Extract indicator values from known column positions
        for ind_id in INDICATOR_MAP:
            val = _extract_value(row_vals, headers, ind_id)
            if val is not None:
                result[ind_id][ent_code] = round(val, 2)

    wb.close()
    logger.info(f"{SOURCE_NAME}: parsed employed_population for {len(result.get('employed_population', {}))} states")
    return result


def _find_entity_code(row_vals: list[str], headers: dict[str, int]) -> str | None:
    """Try to find 2-digit entity code from known header patterns."""
    for key, col in headers.items():
        if col < len(row_vals) and any(
            kw in key for kw in ["cve", "clave", "código", "codigo", "ent", "edo"]
        ):
            val = row_vals[col].strip()
            if val.isdigit() and len(val) in (1, 2):
                return val.zfill(2)
    return None


def _extract_value(row_vals: list[str], headers: dict[str, str], ind_id: str) -> float | None:
    """Extract an indicator value from a row using column name matching."""
    # Try to find relevant columns by keyword matching
    keywords = {
        "employed_population": ["ocup", "pea", "empleo", "tasa ocupacion"],
        "female_employment": ["mujer", "femenin", "female"],
        "hours_worked": ["horas", "hours", "jornada"],
        "remuneration_level": ["ingreso", "remuner", "salario", "income", "wage"],
    }

    for key, col_idx in headers.items():
        if isinstance(col_idx, int):
            kw_list = keywords.get(ind_id, [])
            if any(kw in key.lower() for kw in kw_list):
                val = row_vals[col_idx] if col_idx < len(row_vals) else ""
                try:
                    return float(val.replace(",", "").replace("%", "").replace("$", ""))
                except (ValueError, TypeError):
                    pass
    return None


def _resolve_entity_name_from_row(row_vals: list[str]) -> str | None:
    """Try to find entity name in row and map to code."""
    _name_map = {
        "aguascalientes": "01", "baja california": "02",
        "baja california sur": "03", "campeche": "04",
        "coahuila": "05", "colima": "06", "chiapas": "07",
        "chihuahua": "08", "ciudad de méxico": "09", "distrito federal": "09",
        "durango": "10", "guanajuato": "11", "guerrero": "12",
        "hidalgo": "13", "jalisc": "14", "méxico": "15", "mexico": "15",
        "michoacán": "16", "michoacan": "16", "morelos": "17",
        "nayarit": "18", "nuevo león": "19", "nuevo leon": "19",
        "oaxaca": "20", "puebla": "21", "querétaro": "22",
        "quintana roo": "23", "san luis potosí": "24", "san luis potosi": "24",
        "sinaloa": "25", "sonora": "26", "tabasco": "27",
        "tamaulipas": "28", "tlaxcala": "29", "veracruz": "30",
        "yucatán": "31", "yucatan": "31", "zacatecas": "32",
    }
    for val in row_vals:
        val_lower = val.strip().lower()
        for name, code in _name_map.items():
            if name in val_lower:
                return code
    return None


def get_state_aggregates(
    enoe_data: dict[str, dict[str, float]],
    indicator_id: str,
) -> dict[str, float]:
    """Return state-level values for an indicator."""
    return enoe_data.get(indicator_id, {})


def check_available() -> bool:
    xlsx_files = _find_xlsx_files(CACHE_DIR)
    return len(xlsx_files) > 0
